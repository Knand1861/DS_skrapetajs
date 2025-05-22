import requests
import time
from bs4 import BeautifulSoup
import os
import re
from openpyxl import load_workbook, Workbook
ITEM = "RTX 3060"
def main():
    global ITEM 
    ITEM = input("Enter item name:")
    items = Items()
    # Find items on the web
    items.append(fetch_shop("RD Electronics"))
    items.append(fetch_shop("Euronics"))
    # Sort by price
    items.sort()
    # Save and open excel
    save_to_excel(items)

def save_to_excel(items):
    file_path:str = "data.xlsx"

    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
    else:
        workbook = Workbook()
    # Delete old sheets
    for i in workbook.sheetnames:
        workbook.remove(workbook[i])
    # Add new empty sheet
    workbook.create_sheet("Sheet 1")
    sheet = workbook.active

    print("\nOverall item count:",items.length)

    # Build the Excel
    sheet[f"A1"] = "Item name"
    sheet[f"B1"] = "Price"
    sheet[f"C1"] = "Shop"
    bias:int = 1
    for i in range(1,items.length):
        sheet[f"A{i+bias}"] = items.get(i).name
        sheet[f"B{i+bias}"] = items.get(i).price
        sheet[f"C{i+bias}"] = items.get(i).shop
    workbook.save(file_path)
    workbook.close() 



class Item:
    def __init__(self, name: str, price: str, shop:str):
        self.name:str = name
        self.price:str = price
        self.shop:str = shop

class Items:
    def __init__(self):
        self.items = []
        self.length:int = 0

    def add(self, name: str, price: str, shop:str):
        self.items.append(Item(name, price, shop))
        self.length+=1
    def append(self, items):
        for i in range(0,items.length):
            self.items.append(items.get(i))
            self.length += 1
    def get(self, index:int)->Item:
        return self.items[index]
    def sort(self):
        for i in range(0,self.length):
            minimum = i
            for j in range(i,self.length):
                if self.items[minimum].price > self.items[j].price:
                    minimum = j
            temp = self.items[i]
            self.items[i] = self.items[minimum]
            self.items[minimum] = temp


def fetch_shop(shop:str):
    items = Items()
    print("Looking for",ITEM,"in",shop)
    match shop:
        case "RD Electronics":
            URL = "https://www.rdveikals.lv/search/lv/word/"+ITEM+"/page/1/"
            response = requests.get(URL)
            html:str = response.text
            soup:str = BeautifulSoup(html, "html.parser")
            itemContainers = soup.find_all("li", class_="product")
            for i in itemContainers:
                price = float(re.sub(r"[:a-zA-Z\s]+","",i.find("p", class_="price").get_text().replace("€", "").replace(",", ".").strip()))
                name = i.find("div", class_="product__info__part").find("a").get_text()
                name = " ".join(name.split())
                items.add(name,price,"RD Electronics")
        case "Euronics":
            URL = "https://www.euronics.lv/en/search/"+ITEM
            response = requests.get(URL)
            html:str = response.text
            soup:str = BeautifulSoup(html, "html.parser")
            itemContainers = soup.find_all("article", class_="product-card")
            for i in itemContainers:
                price = i.find("div", class_="price").get_text().split("€")[0]
                price = float(re.sub(r"[:a-zA-Z\s]+","",price.replace("€", "").replace(",", ".").strip()))
                name = i.find("a", class_="product_name").get_text()
                name = " ".join(name.split())
                items.add(name,price,"Euronics")
    
    print("\nShop:",shop,"\nItem count:",items.length)
    return items





if __name__ == "__main__":
    main()